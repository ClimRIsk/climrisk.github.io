"""Generate a blank Excel client intake template with validation and formatting."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


def generate_template(output_path: str | Path) -> None:
    """
    Generate a blank Excel (.xlsx) client intake template with three sheets:
    - Company: one row per company
    - Assets: one row per asset
    - Readme: instructions and valid values
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets
    company_sheet = wb.create_sheet("Company", 0)
    assets_sheet = wb.create_sheet("Assets", 1)
    readme_sheet = wb.create_sheet("Readme", 2)

    # Define header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # =========================================================================
    # COMPANY SHEET
    # =========================================================================
    company_headers = [
        "company_id",
        "company_name",
        "sector",
        "hq_region",
        "total_debt_musd",
        "cash_musd",
        "shares_outstanding_m",
        "current_share_price",
        "revenue_musd",
        "ebitda_musd",
    ]

    for col_idx, header in enumerate(company_headers, start=1):
        cell = company_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Set column widths for Company sheet
    col_widths_company = {
        "A": 15,  # company_id
        "B": 25,  # company_name
        "C": 15,  # sector
        "D": 12,  # hq_region
        "E": 15,  # total_debt_musd
        "F": 12,  # cash_musd
        "G": 18,  # shares_outstanding_m
        "H": 18,  # current_share_price
        "I": 15,  # revenue_musd
        "J": 15,  # ebitda_musd
    }
    for col_letter, width in col_widths_company.items():
        company_sheet.column_dimensions[col_letter].width = width

    # Add region dropdown to Company sheet hq_region column
    region_list = (
        "AU-WA,AU-SA,AU-QLD,AU-NSW,AU-VIC,AU-TAS,AU-NT,US-TX,US-OK,"
        "CL-02,CL-15,CA-QC,CA-AB,NL-NH,GB-ENG,MN-01,global"
    )
    region_dv = DataValidation(type="list", formula1=f'"{region_list}"', allow_blank=True)
    region_dv.error = "Please select a valid region code"
    region_dv.errorTitle = "Invalid Region"
    company_sheet.add_data_validation(region_dv)
    region_dv.add(f"D2:D1000")

    # =========================================================================
    # ASSETS SHEET
    # =========================================================================
    asset_headers = [
        "company_id",
        "asset_id",
        "asset_name",
        "commodity",
        "region",
        "latitude",
        "longitude",
        "baseline_production",
        "production_unit",
        "baseline_unit_cost",
        "energy_cost_share",
        "carrying_value_musd",
        "remaining_life_years",
        "scope1_intensity",
        "scope2_intensity",
        "scope3_intensity",
        "carbon_price_coverage",
        "free_allocation",
    ]

    for col_idx, header in enumerate(asset_headers, start=1):
        cell = assets_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Set column widths for Assets sheet
    col_widths_assets = {
        "A": 15,  # company_id
        "B": 15,  # asset_id
        "C": 20,  # asset_name
        "D": 18,  # commodity
        "E": 12,  # region
        "F": 12,  # latitude
        "G": 12,  # longitude
        "H": 18,  # baseline_production
        "I": 16,  # production_unit
        "J": 18,  # baseline_unit_cost
        "K": 18,  # energy_cost_share
        "L": 18,  # carrying_value_musd
        "M": 18,  # remaining_life_years
        "N": 18,  # scope1_intensity
        "O": 18,  # scope2_intensity
        "P": 18,  # scope3_intensity
        "Q": 20,  # carbon_price_coverage
        "R": 16,  # free_allocation
    }
    for col_letter, width in col_widths_assets.items():
        assets_sheet.column_dimensions[col_letter].width = width

    # Add commodity dropdown
    commodity_list = (
        "iron_ore,copper,aluminium,coal_thermal,coal_metallurgical,"
        "crude_oil,natural_gas,refined_products,cement,electricity"
    )
    commodity_dv = DataValidation(type="list", formula1=f'"{commodity_list}"', allow_blank=True)
    commodity_dv.error = "Please select a valid commodity"
    commodity_dv.errorTitle = "Invalid Commodity"
    assets_sheet.add_data_validation(commodity_dv)
    commodity_dv.add(f"D2:D1000")

    # Add region dropdown
    region_dv_assets = DataValidation(type="list", formula1=f'"{region_list}"', allow_blank=True)
    region_dv_assets.error = "Please select a valid region code"
    region_dv_assets.errorTitle = "Invalid Region"
    assets_sheet.add_data_validation(region_dv_assets)
    region_dv_assets.add(f"E2:E1000")

    # =========================================================================
    # README SHEET
    # =========================================================================
    readme_sheet.column_dimensions["A"].width = 25
    readme_sheet.column_dimensions["B"].width = 80

    row = 1

    # Title
    title_cell = readme_sheet.cell(row=row, column=1, value="CRI Client Intake Template")
    title_cell.font = Font(bold=True, size=14)
    row += 1

    # Instructions section
    readme_sheet.cell(row=row, column=1, value="Instructions")
    readme_sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    instructions = [
        "Complete the Company sheet with one row per company.",
        "Complete the Assets sheet with one row per asset.",
        "All fields are required unless marked as optional.",
        "Use dropdown menus for Commodity and Region fields.",
        "Submit the completed file to the CRI intake system.",
    ]
    for instruction in instructions:
        readme_sheet.cell(row=row, column=1, value="•")
        readme_sheet.cell(row=row, column=2, value=instruction)
        row += 1

    row += 1

    # Company Sheet Fields
    readme_sheet.cell(row=row, column=1, value="Company Sheet Fields")
    readme_sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    company_field_docs = [
        ("company_id", "Unique slug identifier for the company (e.g., bhp_001)"),
        ("company_name", "Full legal name of the company"),
        ("sector", "Industry sector (e.g., Mining, Oil & Gas)"),
        ("hq_region", "Headquarters region (ISO region code, e.g., AU-WA)"),
        ("total_debt_musd", "Total debt in USD millions"),
        ("cash_musd", "Cash and equivalents in USD millions"),
        ("shares_outstanding_m", "Shares outstanding in millions"),
        ("current_share_price", "Current share price in USD"),
        ("revenue_musd", "Latest annual revenue in USD millions"),
        ("ebitda_musd", "Latest annual EBITDA in USD millions"),
    ]
    for field, description in company_field_docs:
        readme_sheet.cell(row=row, column=1, value=field).font = Font(bold=True)
        readme_sheet.cell(row=row, column=2, value=description)
        row += 1

    row += 1

    # Asset Sheet Fields
    readme_sheet.cell(row=row, column=1, value="Assets Sheet Fields")
    readme_sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    asset_field_docs = [
        ("company_id", "Links to the company (must match Company sheet)"),
        ("asset_id", "Unique slug for the asset (e.g., pilbara_001)"),
        ("asset_name", "Human-readable asset name"),
        ("commodity", "Commodity type (use dropdown)"),
        ("region", "Asset region (ISO region code, use dropdown)"),
        ("latitude", "Latitude for hazard lookup (decimal degrees)"),
        ("longitude", "Longitude for hazard lookup (decimal degrees)"),
        ("baseline_production", "Annual production at baseline year"),
        ("production_unit", "Unit of production (Mt, kt, bbl/d, MMboe, MW, etc.)"),
        ("baseline_unit_cost", "USD per production unit"),
        ("energy_cost_share", "Fraction of unit cost from energy (0-1)"),
        ("carrying_value_musd", "Book value in USD millions"),
        ("remaining_life_years", "Remaining mine/asset life in years"),
        ("scope1_intensity", "tCO2e per production unit (direct emissions)"),
        ("scope2_intensity", "tCO2e per production unit (purchased electricity)"),
        ("scope3_intensity", "tCO2e per production unit (value chain)"),
        ("carbon_price_coverage", "Fraction subject to carbon pricing (0-1)"),
        ("free_allocation", "Fraction of emissions with free allowance (0-1)"),
    ]
    for field, description in asset_field_docs:
        readme_sheet.cell(row=row, column=1, value=field).font = Font(bold=True)
        readme_sheet.cell(row=row, column=2, value=description)
        row += 1

    row += 1

    # Valid Commodity Values
    readme_sheet.cell(row=row, column=1, value="Valid Commodity Values")
    readme_sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    commodities = [
        "iron_ore",
        "copper",
        "aluminium",
        "coal_thermal",
        "coal_metallurgical",
        "crude_oil",
        "natural_gas",
        "refined_products",
        "cement",
        "electricity",
    ]
    for commodity in commodities:
        readme_sheet.cell(row=row, column=2, value=commodity)
        row += 1

    row += 1

    # Valid Region Codes
    readme_sheet.cell(row=row, column=1, value="Top 20 Region Codes")
    readme_sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    regions = [
        "AU-WA (Western Australia)",
        "AU-SA (South Australia)",
        "AU-QLD (Queensland)",
        "AU-NSW (New South Wales)",
        "AU-VIC (Victoria)",
        "AU-TAS (Tasmania)",
        "AU-NT (Northern Territory)",
        "US-TX (Texas)",
        "US-OK (Oklahoma)",
        "CL-02 (Antofagasta, Chile)",
        "CL-15 (Arica y Parinacota, Chile)",
        "CA-QC (Quebec, Canada)",
        "CA-AB (Alberta, Canada)",
        "NL-NH (North Holland, Netherlands)",
        "GB-ENG (England, UK)",
        "MN-01 (Arkhangai, Mongolia)",
        "global (Global/Unspecified)",
    ]
    for region in regions:
        readme_sheet.cell(row=row, column=2, value=region)
        row += 1

    # Save workbook
    wb.save(output_path)
    print(f"Template generated: {output_path}")
