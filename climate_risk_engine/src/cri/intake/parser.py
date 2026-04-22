"""Parse a filled-in client intake Excel file into Company objects."""

from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from cri.data.schemas import (
    Asset,
    Commodity,
    Company,
    EmissionsProfile,
    Financials,
)


# Mapping of common aliases to canonical commodity names
COMMODITY_ALIASES = {
    "iron ore": "iron_ore",
    "iron_ore": "iron_ore",
    "copper": "copper",
    "aluminium": "aluminium",
    "aluminum": "aluminium",
    "coal thermal": "coal_thermal",
    "coal_thermal": "coal_thermal",
    "thermal coal": "coal_thermal",
    "coal metallurgical": "coal_metallurgical",
    "coal_metallurgical": "coal_metallurgical",
    "metallurgical coal": "coal_metallurgical",
    "crude oil": "crude_oil",
    "crude_oil": "crude_oil",
    "natural gas": "natural_gas",
    "natural_gas": "natural_gas",
    "refined products": "refined_products",
    "refined_products": "refined_products",
    "cement": "cement",
    "electricity": "electricity",
}


def _normalize_commodity(value: Optional[str]) -> str:
    """
    Normalize commodity string to canonical form.
    Raises ValueError if not recognized.
    """
    if not value:
        raise ValueError("Commodity cannot be empty")

    normalized = value.strip().lower()
    if normalized in COMMODITY_ALIASES:
        return COMMODITY_ALIASES[normalized]

    raise ValueError(
        f"Unknown commodity: {value!r}. Valid values: "
        f"{', '.join(sorted(set(COMMODITY_ALIASES.values())))}"
    )


def _read_sheet(wb, sheet_name: str):
    """Read worksheet and return list of dicts (excluding header row)."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in workbook")

    ws = wb[sheet_name]

    # Read headers from first row
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip())

    if not headers:
        raise ValueError(f"Sheet {sheet_name!r} has no headers")

    # Read data rows
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Skip completely empty rows
        if all(cell.value is None for cell in row):
            continue

        row_data = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                # Get the value, handling different types
                value = cell.value
                row_data[headers[col_idx]] = value

        rows.append((row_idx, row_data))

    return headers, rows


def parse_excel(path: str | Path) -> list[Company]:
    """
    Parse a client intake Excel workbook and return a list of Company objects.

    Raises ValueError with clear row/column information on validation errors.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = load_workbook(path, data_only=True)

    # Read Company sheet
    company_headers, company_rows = _read_sheet(wb, "Company")

    # Read Assets sheet
    asset_headers, asset_rows = _read_sheet(wb, "Assets")

    # Parse companies
    companies_by_id = {}

    for row_idx, company_data in company_rows:
        try:
            company_id_raw = company_data.get("company_id")
            if company_id_raw is None:
                raise ValueError("company_id is required")
            company_id = str(company_id_raw).strip()
            if not company_id:
                raise ValueError("company_id is required")

            company_name = str(company_data.get("company_name", "")).strip()
            if not company_name:
                raise ValueError("company_name is required")

            sector = str(company_data.get("sector", "")).strip()
            if not sector:
                raise ValueError("sector is required")

            hq_region = str(company_data.get("hq_region", "global")).strip() or "global"

            # Parse financial data
            total_debt = float(company_data.get("total_debt_musd") or 0.0)
            cash = float(company_data.get("cash_musd") or 0.0)
            shares_outstanding = float(company_data.get("shares_outstanding_m") or 1.0)
            share_price = float(company_data.get("current_share_price") or 0.0)
            revenue = float(company_data.get("revenue_musd") or 0.0)
            ebitda = float(company_data.get("ebitda_musd") or 0.0)

            # Calculate net_debt and market_cap
            net_debt = total_debt - cash
            market_cap = shares_outstanding * share_price if share_price > 0 else None

            financials = Financials(
                revenue=revenue,
                ebitda=ebitda,
                capex=0.0,  # Not in template
                net_debt=net_debt,
                shares_outstanding=shares_outstanding,
                market_cap=market_cap,
            )

            company = Company(
                id=company_id,
                name=company_name,
                sector=sector,
                hq_region=hq_region,
                financials=financials,
            )
            companies_by_id[company_id] = company

        except (ValueError, TypeError) as e:
            raise ValueError(f"Company sheet row {row_idx}: {e}")

    # Parse assets and attach to companies
    for row_idx, asset_data in asset_rows:
        try:
            company_id = str(asset_data.get("company_id", "")).strip()
            if not company_id:
                raise ValueError("company_id is required")

            if company_id not in companies_by_id:
                raise ValueError(f"company_id {company_id!r} not found in Company sheet")

            asset_id = str(asset_data.get("asset_id", "")).strip()
            if not asset_id:
                raise ValueError("asset_id is required")

            asset_name = str(asset_data.get("asset_name", "")).strip()
            if not asset_name:
                raise ValueError("asset_name is required")

            # Normalize commodity
            commodity_str = asset_data.get("commodity", "")
            commodity_str = _normalize_commodity(commodity_str)
            commodity = Commodity(commodity_str)

            region = str(asset_data.get("region", "global")).strip() or "global"

            latitude = float(asset_data.get("latitude") or 0.0)
            longitude = float(asset_data.get("longitude") or 0.0)

            baseline_production = float(asset_data.get("baseline_production") or 0.0)
            production_unit = str(asset_data.get("production_unit", "tonnes")).strip() or "tonnes"
            baseline_unit_cost = float(asset_data.get("baseline_unit_cost") or 0.0)
            energy_cost_share = float(asset_data.get("energy_cost_share") or 0.3)
            carrying_value = float(asset_data.get("carrying_value_musd") or 0.0)
            remaining_life = asset_data.get("remaining_life_years")
            if remaining_life is not None:
                remaining_life = int(remaining_life)

            scope1 = float(asset_data.get("scope1_intensity") or 0.0)
            scope2 = float(asset_data.get("scope2_intensity") or 0.0)
            scope3 = float(asset_data.get("scope3_intensity") or 0.0)
            carbon_coverage = float(asset_data.get("carbon_price_coverage") or 1.0)
            free_alloc = float(asset_data.get("free_allocation") or 0.0)

            emissions = EmissionsProfile(
                scope1_intensity=scope1,
                scope2_intensity=scope2,
                scope3_intensity=scope3,
                carbon_price_coverage=carbon_coverage,
                free_allocation=free_alloc,
            )

            asset = Asset(
                id=asset_id,
                name=asset_name,
                commodity=commodity,
                region=region,
                baseline_production=baseline_production,
                production_unit=production_unit,
                emissions=emissions,
                carrying_value=carrying_value,
                remaining_life_years=remaining_life,
                baseline_unit_cost=baseline_unit_cost,
                energy_cost_share=energy_cost_share,
            )

            companies_by_id[company_id].assets.append(asset)

        except (ValueError, TypeError) as e:
            raise ValueError(f"Assets sheet row {row_idx}: {e}")

    return list(companies_by_id.values())
